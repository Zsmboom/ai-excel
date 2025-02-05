import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import random

def create_sudoku_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数独游戏"
    
    # 设置列宽和行高
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 8
    for row in range(1, 10):
        ws.row_dimensions[row].height = 40

    # 创建9x9网格
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    # 设置单元格格式
    for row in range(1, 10):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=14, bold=True)
            
            # 设置3x3方格的边框
            if row in [1, 4, 7] and col in [1, 4, 7]:
                cell.border = thick_border
            else:
                cell.border = thin_border

    # 添加数据验证（只允许1-9的数字）
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=9,
        errorStyle="stop",
        errorTitle="输入错误",
        error="请输入1到9之间的数字",
        prompt="输入提示",
        promptTitle="有效输入",
    )
    ws.add_data_validation(dv)
    dv.add("A1:I9")

    # 添加标题和说明
    ws['K1'] = "Excel 数独游戏"
    ws['K1'].font = Font(size=16, bold=True)
    ws['K3'] = "游戏说明："
    ws['K4'] = "1. 在空格中填入1-9的数字"
    ws['K5'] = "2. 每行、每列和每个3x3方格中"
    ws['K6'] = "   的数字不能重复"
    ws['K7'] = "3. 填错时单元格会显示红色"
    ws['K8'] = "4. 全部填对时会显示提示"

    # 添加难度选择
    ws['K10'] = "选择难度："
    ws['K11'] = "初级"
    ws['K12'] = "中级"
    ws['K13'] = "高级"

    # 添加新游戏按钮
    ws['K15'] = "点击VBA按钮开始新游戏"

    # 保存文件
    wb.save('public/games/excel-sudoku.xlsx')

if __name__ == "__main__":
    create_sudoku_workbook() 